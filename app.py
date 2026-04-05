# ✅ FINAL CONSOLIDATED STREAMLIT APP
# ✅ Includes: Split • Merge • Processor • Online Google Sheet Doctor IDs • PDF Tools
# ✅ Fully merged from all 4 parts + corrected + structured
# ✅ Ready to run immediately

import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from zipfile import ZipFile
import re
import os
import base64
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from PIL import Image

try:
    from streamlit_lottie import st_lottie
except:
    st_lottie = None

# =============================================================
# ✅ GOOGLE SHEET: Doctor IDs (REPLACES ID FILE UPLOAD)
# =============================================================
GDRIVE_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/1-u3cegWgrsoXvJYWVwQQRJbyYbdYtjIMDIifnalwHqo/"
    "export?format=xlsx"
)

def load_online_doctor_ids():
    try:
        response = requests.get(GDRIVE_SHEET_URL)
        if response.status_code != 200:
            return {}, "⚠️ Cannot access Google Sheet."

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        headers = [str(ws.cell(1, col).value).strip().lower() if ws.cell(1, col).value else ""
                   for col in range(1, ws.max_column + 1)]

        doctor_col = None
        id_col = None

        for i, h in enumerate(headers):
            if any(x in h for x in ["doctor", "اسم", "name", "دكتور"]):
                doctor_col = i + 1
            if any(x in h for x in ["id", "رقم", "بطاقة", "identity", "national"]):
                id_col = i + 1

        if not doctor_col or not id_col:
            return {}, f"⚠️ Missing required columns: {headers}"

        id_dict = {}
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, doctor_col).value
            docid = ws.cell(row, id_col).value
            if name and docid:
                clean = str(name).strip()
                id_dict[clean] = str(docid).strip()
                id_dict[clean.lower()] = str(docid).strip()
                id_dict[clean.replace(" ", "")] = str(docid).strip()

        return id_dict, f"✅ Loaded {len(id_dict)//3} doctor IDs from online sheet"

    except Exception as e:
        return {}, f"❌ Error: {e}"

# Load once
id_dict, id_message = load_online_doctor_ids()
st.info(id_message)

# =============================================================
# ✅ Helper Functions
# =============================================================
def _safe_name(s):
    return re.sub(r"[^A-Za-z0-9_-]+", "_", str(s))


def display_uploaded_files(files):
    if files:
        for i, f in enumerate(files):
            st.caption(f"{i+1}. {f.name} — {f.size//1024} KB")


def copy_cell_style(src, dst):
    if not src.has_style:
        return
    try:
        if src.font: dst.font = src.font
        if src.fill: dst.fill = src.fill
        if src.border: dst.border = src.border
        if src.alignment: dst.alignment = src.alignment
        dst.number_format = src.number_format
    except:
        pass


def copy_column_widths(src_ws, dst_ws):
    for col in src_ws.column_dimensions:
        width = src_ws.column_dimensions[col].width
        if width:
            dst_ws.column_dimensions[col].width = width

# =============================================================
# ✅ Page Config & UI
# =============================================================
st.set_page_config(
    page_title="Tricks For Excel Tools",
    page_icon="📊",
    layout="wide",
)

st.markdown("## Tricks For Excel — Full Toolkit")
st.markdown("Split • Merge • Processor • Google Sheet IDs • Images → PDF")

# =============================================================
# ✅ Split Tool
# =============================================================
st.markdown("---")
st.markdown("### ✂️ Split Excel/CSV File")

uploaded_file = st.file_uploader(
    "📂 Upload Excel or CSV", type=["xlsx", "csv"],
    key="splitter", accept_multiple_files=False
)

if uploaded_file:
    display_uploaded_files([uploaded_file])

    file_ext = uploaded_file.name.split(".")[-1].lower()

    if file_ext == "csv":
        df = pd.read_csv(uploaded_file)
        st.dataframe(df.head())
    else:
        wb = load_workbook(uploaded_file, data_only=False)
        sheet = st.selectbox("Select sheet", wb.sheetnames)
        df = pd.read_excel(uploaded_file, sheet_name=sheet)
        st.dataframe(df.head())

    col = st.selectbox("Select column to split by", df.columns)

    if st.button("🚀 Start Split"):
        with st.spinner("Processing..."):
            unique_vals = df[col].dropna().unique()
            zip_buf = BytesIO()
            with ZipFile(zip_buf, "w") as z:
                for v in unique_vals:
                    subset = df[df[col] == v]
                    buf = BytesIO()
                    subset.to_csv(buf, index=False, encoding="utf-8-sig")
                    buf.seek(0)
                    z.writestr(f"{_safe_name(v)}.csv", buf.read())

            zip_buf.seek(0)
            st.success("✅ Split complete!")
            st.download_button(
                "⬇️ Download ZIP",
                zip_buf.getvalue(),
                file_name=f"Split_{_safe_name(uploaded_file.name)}.zip",
                mime="application/zip"
            )

# =============================================================
# ✅ Merge Tool
# =============================================================
st.markdown("---")
st.markdown("### 🔁 Merge Excel/CSV Files")

merge_files = st.file_uploader(
    "📂 Upload multiple Excel/CSV files", type=["xlsx", "csv"],
    accept_multiple_files=True, key="merger"
)

if merge_files:
    display_uploaded_files(merge_files)

    if st.button("✨ Merge Files"):
        try:
            excel_only = all(f.name.endswith("xlsx") for f in merge_files)

            if excel_only:
                merged_wb = Workbook()
                merged_ws = merged_wb.active
                merged_ws.title = "Merged_Data"

                row_ptr = 1
                header_done = False

                for file in merge_files:
                    src_wb = load_workbook(file, data_only=False)
                    src_ws = src_wb.active

                    if not header_done:
                        for c in src_ws[1]:
                            dst = merged_ws.cell(row_ptr, c.column, c.value)
                            copy_cell_style(c, dst)
                        row_ptr += 1
                        header_done = True

                    for row in src_ws.iter_rows(min_row=2):
                        for c in row:
                            dst = merged_ws.cell(row_ptr, c.column, c.value)
                            copy_cell_style(c, dst)
                        row_ptr += 1

                # Copy column widths
                first_wb = load_workbook(merge_files[0], data_only=False)
                copy_column_widths(first_wb.active, merged_ws)

                out = BytesIO()
                merged_wb.save(out)
                out.seek(0)

                st.download_button(
                    "⬇️ Download Merged File",
                    out.getvalue(), file_name="Merged.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                dfs = [pd.read_csv(f) if f.name.endswith("csv") else pd.read_excel(f) for f in merge_files]
                df_merged = pd.concat(dfs, ignore_index=True)
                out = BytesIO()
                df_merged.to_excel(out, index=False)
                out.seek(0)

                st.download_button(
                    "⬇️ Download Merged File",
                    out.getvalue(), file_name="Merged.xlsx"
                )

        except Exception as e:
            st.error(f"❌ Error: {e}")

# =============================================================
# ✅ Processor Tool (BUM Update + Google Sheet IDs + CRM reorder)
# =============================================================
st.markdown("---")
st.markdown("### 🧰 Excel Processor Service")

proc_file = st.file_uploader(
    "📂 Upload Excel file to process", type=["xlsx", "xlsm"], key="processor"
)

# Load BUM mapping
BUM_URL = "https://docs.google.com/spreadsheets/d/1XQnQNDFHDKrWYn23ROAeFS2cELNbKurC/export?format=xlsx"

def load_bum_mapping():
    try:
        r = requests.get(BUM_URL)
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        data = []
        headers = [c.value for c in ws[1]]
        mr_idx = headers.index("MR") + 1
        bum_idx = headers.index("BUM") + 1
        for row in range(2, ws.max_row + 1):
            mr = ws.cell(row, mr_idx).value
            bum = ws.cell(row, bum_idx).value
            if mr and bum:
                data.append((str(mr).strip(), str(bum).strip()))
        return dict(data)
    except:
        return {}

bum_dict = load_bum_mapping()

if proc_file:
    st.write("File:", proc_file.name)

    if st.button("⚙️ Start Processing"):
        try:
            wb = load_workbook(proc_file, data_only=False)
            ws = wb.active

            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            header_to_idx = {h: i+1 for i, h in enumerate(headers) if h}

            COLUMN_RENAME_MAP = {
                "L1 Emp Name": "MR",
                "L2 Emp Name": "DM",
                "L3 Emp Name": "AM",
                "L4 Emp Name": "BUM",
            }

            FINAL_COLUMNS = [
                "CRM Interval Date", "Tracking Number", "MR", "DM", "AM",
                "BUM", "Line", "Activity", "Description", "Account Number",
                "Vendor", "Bank", "Cost", "Bricks", "Professionl Accounts",
                "Request Professionals", "Specialities", "Request Date",
            ]

            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Processed_Data"

            # Identify MR, BUM
            mr_col = header_to_idx.get("L1 Emp Name")
            bum_col = header_to_idx.get("L4 Emp Name")

            # Doctor name column
            doc_col = header_to_idx.get("Professionl Accounts")

            # CRM column
            crm_col = None
            for h in headers:
                if h and "CRM Interval Date" in str(h):
                    crm_col = header_to_idx[h]
                    break

            final_map = []

            # CRM first
            if crm_col:
                final_map.append({"name": "CRM Interval Date", "type": "existing", "src": crm_col})
            else:
                final_map.append({"name": "CRM Interval Date", "type": "new", "value": ""})

            # Remaining
            for col in FINAL_COLUMNS[1:]:
                if col == "BUM" and bum_col:
                    final_map.append({"name": "BUM", "type": "bum", "src": bum_col, "mr": mr_col})
                else:
                    found = False
                    for old, new in COLUMN_RENAME_MAP.items():
                        if new == col and old in header_to_idx:
                            final_map.append({"name": col, "type": "existing", "src": header_to_idx[old]})
                            found = True
                            break
                    if not found and col in header_to_idx:
                        final_map.append({"name": col, "type": "existing", "src": header_to_idx[col]})

            # Add ID Number column
            if doc_col:
                final_map.append({"name": "ID Number", "type": "id", "doctor_col": doc_col})

            # Write header row
            for i, col in enumerate(final_map, start=1):
                new_ws.cell(1, i, col["name"])

            matched = 0
            unmatched = []

            # Process rows
            for r in range(2, ws.max_row + 1):
                for c, col in enumerate(final_map, start=1):
                    if col["type"] == "new":
                        new_ws.cell(r, c, "")

                    elif col["type"] == "existing":
                        new_ws.cell(r, c, ws.cell(r, col["src"]).value)

                    elif col["type"] == "bum":
                        mr_value = ws.cell(r, col["mr"]).value
                        if mr_value and str(mr_value).strip() in bum_dict:
                            new_ws.cell(r, c, bum_dict[str(mr_value).strip()])
                        else:
                            new_ws.cell(r, c, ws.cell(r, col["src"]).value)

                    elif col["type"] == "id":
                        doc = ws.cell(r, col["doctor_col"]).value
                        if doc:
                            clean = str(doc).strip()
                            f = (
                                id_dict.get(clean) or
                                id_dict.get(clean.lower()) or
                                id_dict.get(clean.replace(" ", ""))
                            )
                            if f:
                                matched += 1
                                new_ws.cell(r, c, f)
                            else:
                                new_ws.cell(r, c, "")
                                if clean not in unmatched:
                                    unmatched.append(clean)
                        else:
                            new_ws.cell(r, c, "")

            out = BytesIO()
            new_wb.save(out)
            out.seek(0)

            st.success(f"✅ Completed — {matched} IDs matched")
            st.download_button(
                "⬇️ Download Processed File",
                out.getvalue(), file_name="processed_file.xlsx"
            )

            if unmatched:
                st.warning("Unmatched doctors:")
                for d in unmatched[:20]:
                    st.write("-", d)

        except Exception as e:
            st.error(f"❌ Error: {e}")
            st.exception(e)

# =============================================================
# ✅ Images → PDF Tool
# =============================================================
st.markdown("---")
st.markdown("### 🖼️ Images → PDF Converter")

imgs = st.file_uploader(
    "Upload images", type=["jpg", "jpeg", "png"],
    accept_multiple_files=True, key="imgpdf"
)

if imgs:
    if st.button("🖨️ Create PDF"):
        try:
            pil_imgs = []
            for img in imgs:
                im = Image.open(img)
                if im.mode != "RGB":
                    im = im.convert("RGB")
                pil_imgs.append(im)

            pdf_buf = BytesIO()
            pil_imgs[0].save(pdf_buf, save_all=True, append_images=pil_imgs[1:], format="PDF")
            pdf_buf.seek(0)

            st.download_button(
                "⬇️ Download PDF",
                pdf_buf.getvalue(), file_name="Images.pdf",
                mime="application/pdf"
            )
        except Exception as e:
            st.error(f"❌ Error: {e}")
